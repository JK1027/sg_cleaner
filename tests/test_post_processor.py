import pytest
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill
import os

from app.services.post_processor.unmerge_processor import UnmergeProcessor

@pytest.fixture
def complex_excel_file(tmp_path):
    """
    병합 셀, 수식, 데이터 유효성 검사, 조건부 서식이 포함된 복합 엑셀 파일을 생성합니다.
    """
    file_path = tmp_path / "test_complex.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TestSheet"
    
    # 1. 셀 값 및 수식
    ws['A1'] = "Merge Base"
    ws['B1'] = 10
    ws['B2'] = 20
    ws['C1'] = "=SUM(B1:B2)"
    
    # 2. 병합
    ws.merge_cells('A1:A3')
    
    # 3. 데이터 유효성 검사 (드롭다운)
    dv = DataValidation(type="list", formula1='"Option1,Option2"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(ws['D1'])
    
    # 4. 조건부 서식
    red_fill = PatternFill(start_color='FFEE1111', end_color='FFEE1111', fill_type='solid')
    rule = CellIsRule(operator='greaterThan', formula=['15'], fill=red_fill)
    ws.conditional_formatting.add('B1:B2', rule)
    
    wb.save(file_path)
    wb.close()
    
    return str(file_path)

def test_unmerge_processor_complex_file(complex_excel_file):
    # Given
    processor = UnmergeProcessor()
    
    # When
    processed_file = processor.process(complex_excel_file)
    
    # Then
    wb = openpyxl.load_workbook(processed_file)
    ws = wb.active
    
    # 검증 1: 병합 셀 개수 0
    assert len(ws.merged_cells.ranges) == 0, "병합된 셀이 모두 해제되어야 합니다."
    
    # 검증 2: 해제 전후 값 보존 (Fill Down)
    assert ws['A1'].value == "Merge Base"
    assert ws['A2'].value == "Merge Base"
    assert ws['A3'].value == "Merge Base"
    
    # 검증 3: 수식 보존
    assert ws['C1'].value == "=SUM(B1:B2)", "수식이 파괴되지 않아야 합니다."
    
    # 검증 4: 데이터 유효성 검사 유지
    validations = ws.data_validations.dataValidation
    assert len(validations) > 0, "데이터 유효성 검사가 소실되지 않아야 합니다."
    assert validations[0].type == "list"
    
    # 검증 5: 조건부 서식 유지
    cf_rules = ws.conditional_formatting._cf_rules
    assert len(cf_rules) > 0, "조건부 서식이 소실되지 않아야 합니다."
    
    wb.close()
