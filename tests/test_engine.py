import os
import unittest
import openpyxl
from app.services.detector import AnonymizeDetector
from app.services.excel_service import ExcelService
from app.models.detection_model import DetectionItem
from app.utils.path_helper import get_project_root

class TestAnonymizeEngine(unittest.TestCase):
    """
    익명화 탐지 엔진 및 Excel Safe Save 파이프라인 단위 테스트
    """
    def setUp(self):
        self.project_root = get_project_root()
        self.test_dir = self.project_root / "tests" / "temp_test"
        os.makedirs(self.test_dir, exist_ok=True)
        
        # 1. 테스트용 더미 엑셀 파일 생성
        self.test_excel_path = str(self.test_dir / "test_data.xlsx")
        
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "학급기록"
        
        # 일반 데이터 셀
        ws1["A1"] = "순번"
        ws1["B1"] = "학생 이름"
        ws1["C1"] = "특이사항 및 학교 활동"
        
        ws1["A2"] = 1
        ws1["B2"] = "김민수" # 탐색 대상
        ws1["C2"] = "김민수 학생은 서울중학교 과학 탐구반에서 주도적으로 활동함." # 탐색 대상 (이름, 학교 포함)
        
        # 수식 셀 (스킵되어야 함)
        ws1["A3"] = "=SUM(A2:A2)"
        ws1["B3"] = "이서연" # 탐색 대상
        ws1["C3"] = "이서연 학생은 항상 성실함."
        
        # 병합 셀 구성 (B4:C4 병합)
        ws1.merge_cells("B4:C4")
        ws1["B4"] = "홍길동 학생이 전출 감." # 대표 셀
        
        # 숨김 시트 생성
        ws2 = wb.create_sheet("비공개기록")
        ws2.sheet_state = "hidden" # 숨김 속성 부여
        ws2["A1"] = "비공개"
        ws2["B1"] = "김민수" # 탐지되면 안 됨 (숨김 시트이므로)
        
        wb.save(self.test_excel_path)
        wb.close()

    def tearDown(self):
        # 테스트 임시 파일 및 폴더 삭제 (Clean up)
        if os.path.exists(self.test_excel_path):
            os.remove(self.test_excel_path)
            
        test_out = self.test_dir / "test_data_anonymized.xlsx"
        if os.path.exists(test_out):
            os.remove(test_out)
            
        mapping_out = self.test_dir / "mapping.csv"
        if os.path.exists(mapping_out):
            os.remove(mapping_out)
            
        if os.path.exists(self.test_dir):
            os.rmdir(self.test_dir)

    def test_detector_scanning(self):
        """규칙 기반 탐지 및 병합 셀/숨김 시트/수식 필터링 검증"""
        detector = AnonymizeDetector(
            student_names=["김민수", "이서연", "홍길동"],
            school_names=["서울중학교"]
        )
        
        results = detector.scan_workbook(self.test_excel_path)
        
        # 1. 탐지 개수 확인
        # 김민수: B2, C2 (2개)
        # 이서연: B3 (1개 - A3 수식 셀은 제외)
        # 홍길동: B4 (1개 - 병합 셀 대표 셀인 B4만 검출되고 C4는 MergedCell이므로 스킵됨)
        # 서울중학교: C2 (1개)
        # 비공개기록 시트는 숨김 시트이므로 탐지 안 됨.
        # 총 기대 탐지 개수: 6개
        self.assertEqual(len(results), 6)
        
        # 2. 매핑 대장 자동 인덱싱 생성 확인
        mapping = detector.get_full_mapping()
        self.assertIn("김민수", mapping)
        self.assertEqual(mapping["김민수"], "학생1")
        self.assertEqual(mapping["이서연"], "학생2")
        self.assertEqual(mapping["홍길동"], "학생3")
        self.assertEqual(mapping["서울중학교"], "학교A")

    def test_safe_save_flow(self):
        """Safe Save 4단계 저장 및 데이터 무결성 검증"""
        detector = AnonymizeDetector(
            student_names=["김민수"],
            school_names=["서울중학교"]
        )
        results = detector.scan_workbook(self.test_excel_path)
        
        excel_service = ExcelService()
        output_dir = str(self.test_dir)
        
        # Safe Save 치환 저장 실행
        final_path = excel_service.apply_replacements_safe(self.test_excel_path, results, output_dir)
        
        self.assertTrue(os.path.exists(final_path))
        self.assertIn("test_data_anonymized.xlsx", final_path)
        
        # 저장본이 다시 제대로 열리고 치환이 반영되었는지 확인
        wb = openpyxl.load_workbook(final_path, data_only=True)
        ws = wb["학급기록"]
        
        # 김민수 -> 학생1로 잘 바뀌었는지 확인
        self.assertEqual(ws["B2"].value, "학생1")
        self.assertEqual(ws["C2"].value, "학생1 학생은 학교A 과학 탐구반에서 주도적으로 활동함.")
        
        wb.close()

    def test_korean_path_handling(self):
        """한글 경로 및 파일명이 포함된 파일 처리 능력 검증"""
        # 한글 및 특수문자 파일 경로 설정
        korean_file_path = str(self.test_dir / "생활기록부_테스트_★데이터.xlsx")
        
        # 임시 파일 복사
        import shutil
        shutil.copy2(self.test_excel_path, korean_file_path)
        
        try:
            detector = AnonymizeDetector(student_names=["김민수"], school_names=[])
            results = detector.scan_workbook(korean_file_path)
            
            # 김민수 2건 검출 확인
            self.assertEqual(len(results), 2)
            
            # Safe Save 변환 적용
            excel_service = ExcelService()
            final_path = excel_service.apply_replacements_safe(korean_file_path, results, str(self.test_dir))
            
            self.assertTrue(os.path.exists(final_path))
            self.assertIn("생활기록부_테스트_★데이터_anonymized.xlsx", final_path)
            
            # 클린업
            if os.path.exists(final_path):
                os.remove(final_path)
        finally:
            if os.path.exists(korean_file_path):
                os.remove(korean_file_path)

    def test_missing_file_handling(self):
        """존재하지 않는 파일 처리 시 예외 롤백 처리 검증"""
        excel_service = ExcelService()
        dummy_item = DetectionItem(
            file_path="non_existent.xlsx",
            sheet_name="Sheet1",
            cell_address="A1",
            original_value="김민수",
            match_value="김민수",
            replacement="학생1",
            approved=True
        )
        
        # FileNotFoundError 예외 발생 확인
        with self.assertRaises(FileNotFoundError):
            excel_service.apply_replacements_safe("non_existent.xlsx", [dummy_item], str(self.test_dir))

    def test_delete_mode(self):
        """삭제(제거) 모드 동작 검증 (대체 텍스트 설정 및 빈 값 치환)"""
        # 1. '***' 대체 텍스트 테스트
        detector_star = AnonymizeDetector(
            student_names=[],
            school_names=[],
            delete_keywords=["김민수", "서울중학교"],
            delete_replacement="***"
        )
        results_star = detector_star.scan_workbook(self.test_excel_path)
        
        # 매핑 검증
        mapping_star = detector_star.get_full_mapping()
        self.assertEqual(mapping_star["김민수"], "***")
        self.assertEqual(mapping_star["서울중학교"], "***")
        
        excel_service = ExcelService()
        output_dir = str(self.test_dir)
        final_path_star = excel_service.apply_replacements_safe(self.test_excel_path, results_star, output_dir)
        
        # 파일 내용 검증 (*** 치환 확인)
        wb = openpyxl.load_workbook(final_path_star, data_only=True)
        ws = wb["학급기록"]
        self.assertEqual(ws["B2"].value, "***")
        self.assertEqual(ws["C2"].value, "*** 학생은 *** 과학 탐구반에서 주도적으로 활동함.")
        wb.close()
        os.remove(final_path_star)

        # 2. 빈 문자열('') 대체 텍스트 (완전 삭제) 테스트
        detector_empty = AnonymizeDetector(
            student_names=[],
            school_names=[],
            delete_keywords=["김민수", "서울중학교"],
            delete_replacement=""
        )
        results_empty = detector_empty.scan_workbook(self.test_excel_path)
        
        # 매핑 검증
        mapping_empty = detector_empty.get_full_mapping()
        self.assertEqual(mapping_empty["김민수"], "")
        self.assertEqual(mapping_empty["서울중학교"], "")
        
        final_path_empty = excel_service.apply_replacements_safe(self.test_excel_path, results_empty, output_dir)
        
        # 파일 내용 검증 (공백 제거 확인)
        wb = openpyxl.load_workbook(final_path_empty, data_only=True)
        ws = wb["학급기록"]
        self.assertIn(ws["B2"].value, ("", None)) # 완전히 공백 또는 None
        self.assertEqual(ws["C2"].value, " 학생은  과학 탐구반에서 주도적으로 활동함.") # 단어 빠짐
        wb.close()
        os.remove(final_path_empty)

    def test_keyword_conflict_and_order_determinism(self):
        """키워드 충돌 시 우선순위 필터링 및 탐색 순서 결정성 검증"""
        from app.models.app_state import AppState
        from app.controllers.app_controller import AppController
        
        state = AppState()
        controller = AppController(state)
        
        # 1. 중복/충돌 키워드 입력
        # '김민수'가 학생명, 학교명, 삭제 단어에 모두 존재
        # '서울중학교'가 학교명, 삭제 단어에 존재
        students = ["김민수", "이서연", "김민수"]
        schools = ["서울중학교", "김민수", "서울중학교"]
        deletes = ["김민수", "서울중학교", "삭제어1"]
        
        controller.update_input_patterns(students, schools, deletes)
        
        # 중복 제거 및 우선순위 필터링 검증
        # 학생명: ['김민수', '이서연']
        # 학교명: ['서울중학교'] ('김민수'는 학생명이므로 제거됨)
        # 삭제어: ['삭제어1'] ('김민수', '서울중학교'는 학생명/학교명이므로 제거됨)
        self.assertEqual(state.student_names, ["김민수", "이서연"])
        self.assertEqual(state.school_names, ["서울중학교"])
        self.assertEqual(state.delete_keywords, ["삭제어1"])
        
        # 2. 결정성(Determinism) 검증 (입력 순서가 매핑 인덱스 순서에 영향을 미치는지 확인)
        detector = AnonymizeDetector(
            student_names=["김민수", "이서연"],
            school_names=["서울중학교"]
        )
        detector.scan_workbook(self.test_excel_path)
        mapping = detector.get_full_mapping()
        
        # 순서가 뒤집히지 않고 입력된 순서대로 인덱싱 부여
        self.assertEqual(mapping["김민수"], "학생1")
        self.assertEqual(mapping["이서연"], "학생2")

    def test_substring_overlap_prevention(self):
        """동일 셀 내 부분 문자열 중복 매칭 차단 알고리즘 검증"""
        # '김민수'와 '민수'가 동시에 패턴으로 등록됨
        detector = AnonymizeDetector(
            student_names=["김민수", "민수"],
            school_names=[]
        )
        
        results = detector.scan_workbook(self.test_excel_path)
        
        # '민수'에 대한 개별 매칭이 중복 생성되었는지 검증 (김민수만 매칭되어야 함)
        match_values = [item.match_value for item in results]
        self.assertIn("김민수", match_values)
        self.assertNotIn("민수", match_values) # 민수는 김민수의 부분 문자열이므로 오버랩 제외 확인

    def test_worker_cancellation(self):
        """QThread 백그라운드 Worker 스레드 작업 취소 및 안전 재구동 검증"""
        from app.services.worker import DetectionWorker
        
        worker = DetectionWorker(
            file_paths=[self.test_excel_path],
            student_names=["김민수"],
            school_names=["서울중학교"]
        )
        
        # 작업 취소 설정
        worker.cancel()
        
        # 취소 이벤트가 수신 및 기록되어 finished 대신 error_occurred가 불리는지 확인하는 모의 슬롯
        error_called = []
        def on_error(msg):
            error_called.append(msg)
            
        worker.error_occurred.connect(on_error)
        worker.run() # 동기식 run() 구동으로 취소 플래그 인터셉트 동작 테스트
        
        self.assertTrue(len(error_called) > 0)
        self.assertIn("취소", error_called[0])

if __name__ == "__main__":
    unittest.main()
