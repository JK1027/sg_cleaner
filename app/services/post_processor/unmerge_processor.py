import openpyxl
from .base_processor import BaseProcessor
from app.utils.logger import logger

class UnmergeProcessor(BaseProcessor):
    id: str = "unmerge"
    name: str = "셀 병합 해제 및 데이터 자동 채우기"
    description: str = "병합 셀을 해제하고 값을 채웁니다. (생기부 점검용 파일 생성 시 권장)"

    def process(self, file_path: str) -> str:
        """
        주어진 엑셀 파일의 모든 시트에서 병합된 셀을 해제하고 원래 값을 채워 넣은 뒤
        동일한 파일 이름으로 저장합니다.
        """
        logger.info(f"[UnmergeProcessor] 시작: {file_path}")
        wb = openpyxl.load_workbook(file_path)
        
        total_unmerged = self._unmerge_all_and_preserve(wb, fill_down=True, remove_empty_rows=False)
        
        # 파일 저장 (기존 파일 덮어쓰기)
        wb.save(file_path)
        wb.close()
        
        logger.info(f"[UnmergeProcessor] 완료. 해제된 셀 묶음 개수: {total_unmerged}")
        return file_path

    def _unmerge_all_and_preserve(self, workbook: openpyxl.Workbook, fill_down: bool = True, remove_empty_rows: bool = False) -> int:
        """
        워크북 내 모든 시트의 모든 병합 셀을 해제하고 원래 값을 보존하거나 복사합니다.
        (excel_merge_cleaner에서 이관된 핵심 로직)
        """
        total_unmerged = 0
        for sheet in workbook.worksheets:
            logger.info(f"Scanning sheet '{sheet.title}' to unmerge all cells.")
            
            merged_ranges = list(sheet.merged_cells.ranges)
            merged_rows = set()
            for cell_range in merged_ranges:
                for r in range(cell_range.min_row, cell_range.max_row + 1):
                    merged_rows.add(r)
            
            sheet_unmerged = 0
            if merged_ranges:
                for cell_range in merged_ranges:
                    min_row, min_col = cell_range.min_row, cell_range.min_col
                    max_row, max_col = cell_range.max_row, cell_range.max_col
                    
                    top_left = sheet.cell(row=min_row, column=min_col)
                    val = top_left.value
                    top_left_style = top_left._style
                    
                    sheet.unmerge_cells(
                        start_row=min_row,
                        start_column=min_col,
                        end_row=max_row,
                        end_column=max_col
                    )
                    sheet_unmerged += 1
                    
                    # 복원된 일반 Cell 객체들에 좌상단 셀 서식 복제
                    for r in range(min_row, max_row + 1):
                        for c in range(min_col, max_col + 1):
                            sheet.cell(row=r, column=c)._style = top_left_style
                    
                    width = max_col - min_col + 1
                    height = max_row - min_row + 1
                    is_formula = isinstance(val, str) and val.strip().startswith('=')
                    
                    if fill_down and not is_formula and width == 1 and height > 1:
                        # 순수 세로 병합만 Fill Down 진행
                        for r in range(min_row, max_row + 1):
                            sheet.cell(row=r, column=min_col).value = val
                    else:
                        # 좌상단 셀만 값 보존, 나머지는 None (수식 제외)
                        sheet.cell(row=min_row, column=min_col).value = val
                        for r in range(min_row, max_row + 1):
                            for c in range(min_col, max_col + 1):
                                if r == min_row and c == min_col:
                                    continue
                                cell = sheet.cell(row=r, column=c)
                                c_val = cell.value
                                if not (isinstance(c_val, str) and c_val.strip().startswith('=')):
                                    cell.value = None

            # 열 너비 자동 조정
            for col in sheet.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    val = cell.value
                    if val is not None:
                        length = 0
                        for char in str(val):
                            if ord(char) > 255:
                                length += 2
                            else:
                                length += 1
                        if length > max_len:
                            max_len = length
                if max_len > 0:
                    calculated_width = max(max_len + 3, 10)
                    calculated_width = min(calculated_width, 120)
                    sheet.column_dimensions[col_letter].width = calculated_width

            # 행 높이 리셋
            for r_idx in merged_rows:
                sheet.row_dimensions[r_idx].height = None

            # 빈 행 제거 (사용하는 경우)
            if remove_empty_rows:
                deleted_rows_count = 0
                for r_idx in range(sheet.max_row, 0, -1):
                    is_empty = True
                    for col_idx in range(1, sheet.max_column + 1):
                        val = sheet.cell(row=r_idx, column=col_idx).value
                        if val is not None and str(val).strip() != "":
                            is_empty = False
                            break
                    if is_empty:
                        sheet.delete_rows(r_idx, 1)
                        deleted_rows_count += 1
                if deleted_rows_count > 0:
                    logger.info(f"Removed {deleted_rows_count} empty rows from sheet '{sheet.title}'.")

            # 검증
            if len(sheet.merged_cells.ranges) > 0:
                raise ValueError(f"Merged cells still remain in sheet '{sheet.title}' after unmerging.")
                
            logger.info(f"Successfully unmerged {sheet_unmerged} ranges in sheet '{sheet.title}'.")
            total_unmerged += sheet_unmerged
            
        return total_unmerged
