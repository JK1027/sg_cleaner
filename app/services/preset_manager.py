import os
import json
import uuid
from datetime import datetime
from PySide6.QtCore import QStandardPaths
from app.utils.logger import logger

class PresetManager:
    """
    사용자 로컬 AppData 디렉토리 내 presets 폴더에서 프리셋(JSON)의 CRUD를 담당하고
    임시 입력 유실을 방지하는 draft.json 자동 저장을 처리하는 서비스 클래스입니다.
    """
    
    @staticmethod
    def get_presets_dir() -> str:
        """프리셋이 저장되는 AppData 디렉토리 경로를 반환합니다. 디렉토리가 없으면 자동 생성합니다."""
        app_data = QStandardPaths.writableLocation(QStandardPaths.AppDataLocation)
        if not app_data:
            # Fallback to user home directory
            app_data = os.path.join(os.path.expanduser("~"), ".sg_cleaner")
        
        # QStandardPaths가 앱 명칭을 포함하지 않는 경우를 대비해 확실히 sg_cleaner 폴더 분리
        if "sg_cleaner" not in app_data.lower() and "sgcleaner" not in app_data.lower():
            app_data = os.path.join(app_data, "SgCleaner")
            
        presets_dir = os.path.join(app_data, "presets")
        os.makedirs(presets_dir, exist_ok=True)
        return presets_dir

    @classmethod
    def get_presets_info(cls) -> dict[str, dict]:
        """
        저장된 모든 프리셋 파일(preset_*.json)을 스캔하여
        { file_id: { "name": ..., "updated_at": ..., "created_at": ... } } 형식으로 반환합니다.
        수정일(updated_at) 내림차순(최신순)으로 정렬해 반환합니다.
        """
        presets_dir = cls.get_presets_dir()
        presets_info = {}
        
        if not os.path.exists(presets_dir):
            return presets_info
            
        for filename in os.listdir(presets_dir):
            if filename.startswith("preset_") and filename.endswith(".json"):
                file_path = os.path.join(presets_dir, filename)
                file_id = filename[7:-5] # "preset_"과 ".json" 제외
                
                try:
                    with open(file_path, "r", encoding="utf-8") as f:
                        data = json.load(f)
                        presets_info[file_id] = {
                            "name": data.get("name", "이름 없음"),
                            "created_at": data.get("created_at", ""),
                            "updated_at": data.get("updated_at", ""),
                            "file_path": file_path
                        }
                except Exception as e:
                    logger.error(f"프리셋 파일 읽기 실패 ({filename}): {str(e)}")
                    
        # updated_at을 기준으로 정렬하여 딕셔너리 재구성
        sorted_presets = sorted(
            presets_info.items(),
            key=lambda x: x[1].get("updated_at", ""),
            reverse=True
        )
        return dict(sorted_presets)

    @classmethod
    def load_preset(cls, file_id: str) -> dict:
        """지정한 file_id의 프리셋 데이터를 파일로부터 읽어와 딕셔너리로 반환합니다."""
        presets_dir = cls.get_presets_dir()
        file_path = os.path.join(presets_dir, f"preset_{file_id}.json")
        
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"프리셋 파일을 찾을 수 없습니다: {file_path}")
            
        with open(file_path, "r", encoding="utf-8") as f:
            return json.load(f)

    @classmethod
    def create_preset(cls, display_name: str, payload: dict) -> str:
        """
        새로운 고유 UUID 기반의 프리셋을 생성하고 저장합니다.
        생성된 file_id를 반환합니다.
        """
        file_id = uuid.uuid4().hex
        presets_dir = cls.get_presets_dir()
        file_path = os.path.join(presets_dir, f"preset_{file_id}.json")
        
        now = datetime.now().isoformat()
        
        preset_data = {
            "version": "1.0",
            "name": display_name,
            "created_at": now,
            "updated_at": now,
            "students": payload.get("students", []),
            "schools": payload.get("schools", []),
            "delete_keywords": payload.get("delete_keywords", []),
            "delete_replacement": payload.get("delete_replacement", "")
        }
        
        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(preset_data, f, ensure_ascii=False, indent=2)
            
        logger.info(f"새 프리셋 생성 성공: {display_name} ({file_path})")
        return file_id

    @classmethod
    def save_preset(cls, file_id: str, display_name: str, payload: dict) -> None:
        """기존 프리셋 정보를 덮어쓰기 저장합니다."""
        presets_dir = cls.get_presets_dir()
        file_path = os.path.join(presets_dir, f"preset_{file_id}.json")
        
        # 기존 데이터 로드하여 created_at 보존 시도
        created_at = datetime.now().isoformat()
        if os.path.exists(file_path):
            try:
                with open(file_path, "r", encoding="utf-8") as f:
                    old_data = json.load(f)
                    created_at = old_data.get("created_at", created_at)
            except Exception:
                pass
                
        now = datetime.now().isoformat()
        
        preset_data = {
            "version": "1.0",
            "name": display_name,
            "created_at": created_at,
            "updated_at": now,
            "students": payload.get("students", []),
            "schools": payload.get("schools", []),
            "delete_keywords": payload.get("delete_keywords", []),
            "delete_replacement": payload.get("delete_replacement", "")
        }
        
        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(preset_data, f, ensure_ascii=False, indent=2)
            
        logger.info(f"프리셋 저장 완료: {display_name} ({file_path})")

    @classmethod
    def delete_preset(cls, file_id: str) -> None:
        """지정한 프리셋 파일을 물리 삭제합니다."""
        presets_dir = cls.get_presets_dir()
        file_path = os.path.join(presets_dir, f"preset_{file_id}.json")
        
        if os.path.exists(file_path):
            os.remove(file_path)
            logger.info(f"프리셋 파일 삭제 완료: {file_path}")
        else:
            logger.warning(f"삭제하려는 프리셋 파일이 존재하지 않습니다: {file_path}")

    @classmethod
    def save_draft(cls, payload: dict) -> None:
        """현재 입력 폼의 상태를 draft.json으로 임시 저장합니다."""
        app_data = QStandardPaths.writableLocation(QStandardPaths.AppDataLocation)
        if not app_data:
            app_data = os.path.join(os.path.expanduser("~"), ".sg_cleaner")
            
        if "sg_cleaner" not in app_data.lower() and "sgcleaner" not in app_data.lower():
            app_data = os.path.join(app_data, "SgCleaner")
            
        os.makedirs(app_data, exist_ok=True)
        draft_path = os.path.join(app_data, "draft.json")
        
        try:
            with open(draft_path, "w", encoding="utf-8") as f:
                json.dump(payload, f, ensure_ascii=False, indent=2)
            # 디버그 수준 로그만 출력 (너무 빈번한 로깅 차단)
            logger.debug(f"임시 드래프트 자동 저장 성공: {draft_path}")
        except Exception as e:
            logger.error(f"임시 드래프트 저장 실패: {str(e)}")

    @classmethod
    def load_draft(cls) -> dict | None:
        """임시 저장된 draft.json을 읽어와 반환합니다. 없으면 None을 반환합니다."""
        app_data = QStandardPaths.writableLocation(QStandardPaths.AppDataLocation)
        if not app_data:
            app_data = os.path.join(os.path.expanduser("~"), ".sg_cleaner")
            
        if "sg_cleaner" not in app_data.lower() and "sgcleaner" not in app_data.lower():
            app_data = os.path.join(app_data, "SgCleaner")
            
        draft_path = os.path.join(app_data, "draft.json")
        
        if os.path.exists(draft_path):
            try:
                with open(draft_path, "r", encoding="utf-8") as f:
                    return json.load(f)
            except Exception as e:
                logger.error(f"임시 드래프트 로드 중 오류 발생: {str(e)}")
                return None
        return None

    @classmethod
    def import_preset_from_excel(cls, file_path: str) -> dict:
        """엑셀 프리셋 파일을 읽어 표준 프리셋 페이로드 딕셔너리를 반환합니다. (openpyxl 전용)"""
        import openpyxl
        
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        except PermissionError as pe:
            raise PermissionError("파일이 이미 열려 있습니다. 엑셀 프로그램을 종료한 후 다시 시도해 주세요.") from pe
        except Exception as e:
            raise ValueError("지원하지 않거나 손상된 엑셀 파일입니다.") from e

        try:
            ws = wb.active
            
            # 1. 헤더 매핑 식별
            headers = []
            for col in range(1, ws.max_column + 1):
                val = ws.cell(row=1, column=col).value
                headers.append(str(val).strip() if val is not None else "")

            col_indices = {
                "students": None,
                "student_replacements": None,
                "schools": None,
                "school_replacements": None,
                "deletes": None,
                "replacement": None
            }
            
            def normalize(s: str) -> str:
                return s.replace(" ", "").replace("_", "").lower()

            for idx, h in enumerate(headers):
                h_norm = normalize(h)
                if any(x in h_norm for x in ["삭제대체", "대체텍스트", "대체"]):
                    col_indices["replacement"] = idx + 1
                elif any(x in h_norm for x in ["학생변경예정", "학생변경", "학생가명"]):
                    col_indices["student_replacements"] = idx + 1
                elif any(x in h_norm for x in ["학생이름", "학생명", "학생"]):
                    col_indices["students"] = idx + 1
                elif any(x in h_norm for x in ["학교변경예정", "학교변경", "학교가명"]):
                    col_indices["school_replacements"] = idx + 1
                elif any(x in h_norm for x in ["학교명", "학교이름", "학교"]):
                    col_indices["schools"] = idx + 1
                elif any(x in h_norm for x in ["삭제할단어", "삭제단어", "삭제"]):
                    col_indices["deletes"] = idx + 1

            # 유효한 열이 하나도 없는 경우 오류
            if not any(col_indices.values()):
                raise ValueError("올바른 엑셀 양식이 아닙니다. 필수 열(학생 이름, 학교명, 삭제 단어 등)을 찾을 수 없습니다.")

            students = []
            schools = []
            deletes = []
            replacement = ""

            # 2. 데이터 순회 로드 (iter_rows 활용하여 안전하게 순회)
            row_idx = 2
            for row_cells in ws.iter_rows(min_row=2):
                def get_cell_val(col_idx):
                    if col_idx is not None and col_idx <= len(row_cells):
                        return row_cells[col_idx - 1].value
                    return None

                # 학생명 파싱
                students_val = get_cell_val(col_indices["students"])
                if students_val is not None:
                    name_str = str(students_val).strip()
                    if name_str:
                        rep_val = get_cell_val(col_indices["student_replacements"])
                        rep_str = str(rep_val).strip() if rep_val is not None else ""
                        if rep_str:
                            students.append(f"{name_str}:{rep_str}")
                        else:
                            students.append(name_str)

                # 학교명 파싱
                schools_val = get_cell_val(col_indices["schools"])
                if schools_val is not None:
                    school_str = str(schools_val).strip()
                    if school_str:
                        rep_val = get_cell_val(col_indices["school_replacements"])
                        rep_str = str(rep_val).strip() if rep_val is not None else ""
                        if rep_str:
                            schools.append(f"{school_str}:{rep_str}")
                        else:
                            schools.append(school_str)

                # 삭제할 단어 파싱
                deletes_val = get_cell_val(col_indices["deletes"])
                if deletes_val is not None:
                    val_str = str(deletes_val).strip()
                    if val_str:
                        deletes.append(val_str)

                # 삭제 대체 텍스트는 첫 데이터 행(2행)에서만 로드
                if row_idx == 2:
                    rep_val = get_cell_val(col_indices["replacement"])
                    if rep_val is not None:
                        replacement = str(rep_val).strip()
                row_idx += 1

            # 중복 제거 (순서 보존)
            return {
                "students": list(dict.fromkeys(students)),
                "schools": list(dict.fromkeys(schools)),
                "delete_keywords": list(dict.fromkeys(deletes)),
                "delete_replacement": replacement
            }
        finally:
            wb.close()

    @classmethod
    def export_preset_to_excel(cls, payload: dict, file_path: str) -> None:
        """화면 상태 페이로드를 스타일 서식이 적용된 규격 엑셀 파일로 출력합니다."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "익명화 프리셋 양식"

        # 1. 헤더 생성
        headers = ["학생 이름", "학생 변경 예정", "학교명", "학교 변경 예정", "삭제할 단어", "삭제 대체 텍스트"]
        ws.append(headers)

        # 2. 데이터 로드 및 작성
        students = payload.get("students", [])
        schools = payload.get("schools", [])
        deletes = payload.get("delete_keywords", [])
        replacement = payload.get("delete_replacement", "")

        import re
        student_pattern = re.compile(r"^(\d{3,6})\s*([가-힣]{2,5})$")

        # 학생 데이터 파싱
        parsed_students = []
        for s in students:
            if ":" in s:
                parts = s.split(":", 1)
                parsed_students.append((parts[0].strip(), parts[1].strip()))
            else:
                match = student_pattern.match(s)
                if match:
                    parsed_students.append((match.group(2), f"학생{match.group(1)}"))
                else:
                    parsed_students.append((s, ""))

        # 학교 데이터 파싱
        parsed_schools = []
        for sch in schools:
            if ":" in sch:
                parts = sch.split(":", 1)
                parsed_schools.append((parts[0].strip(), parts[1].strip()))
            else:
                parsed_schools.append((sch, ""))

        max_len = max(len(parsed_students), len(parsed_schools), len(deletes))
        
        for i in range(max_len):
            row_data = [
                parsed_students[i][0] if i < len(parsed_students) else "",
                parsed_students[i][1] if i < len(parsed_students) else "",
                parsed_schools[i][0] if i < len(parsed_schools) else "",
                parsed_schools[i][1] if i < len(parsed_schools) else "",
                deletes[i] if i < len(deletes) else "",
                replacement if i == 0 else ""
            ]
            ws.append(row_data)

        # 3. 스타일 서식 가공
        header_font = Font(name="맑은 고딕", size=11, bold=True, color="000000")
        header_fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        
        thin_side = Side(border_style="thin", color="CCCCCC")
        border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        # 헤더 스타일
        for col_idx in range(1, 7):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border

        # 데이터 스타일
        for r in range(2, max_len + 2):
            for col_idx in range(1, 7):
                cell = ws.cell(row=r, column=col_idx)
                cell.border = border
                cell.font = Font(name="맑은 고딕", size=10)
                if col_idx in (2, 4, 6):
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align

        # 열 너비 설정
        col_widths = {"A": 22, "B": 22, "C": 22, "D": 22, "E": 22, "F": 26}
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width

        # 필터 적용
        if max_len > 0:
            ws.auto_filter.ref = f"A1:F{max_len + 1}"

        try:
            wb.save(file_path)
        except PermissionError as pe:
            raise PermissionError("파일이 이미 열려 있습니다. 엑셀 프로그램을 종료한 후 다시 시도해 주세요.") from pe
        except Exception as e:
            raise IOError(f"엑셀 파일 저장 중 오류가 발생했습니다: {str(e)}") from e
        finally:
            wb.close()

    @classmethod
    def convert_neis_excel_to_preset(cls, file_path: str) -> list[str]:
        """나이스 학적현황 엑셀을 파싱하여 ['성명:학생학번', ...] 리스트로 변환합니다."""
        import openpyxl
        
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        except PermissionError as pe:
            raise PermissionError("파일이 이미 열려 있습니다. 엑셀 프로그램을 종료한 후 다시 시도해 주세요.") from pe
        except Exception as e:
            raise ValueError("지원하지 않거나 손상된 엑셀 파일입니다.") from e

        try:
            ws = wb.active
            
            # 1. 헤더 행 스캔
            header_row_idx = None
            col_indices = {"grade": None, "class": None, "number": None, "name": None}
            
            # 상단 30행 이내에서 헤더를 탐색 (보통 제목 행 아래에 있음)
            for r in range(1, min(ws.max_row + 1, 30)):
                row_vals = []
                for c in range(1, ws.max_column + 1):
                    val = ws.cell(row=r, column=c).value
                    row_vals.append(str(val).strip() if val is not None else "")
                
                grade_idx = class_idx = num_idx = name_idx = None
                for idx, val in enumerate(row_vals):
                    val_norm = val.replace(" ", "")
                    if val_norm in ["학년", "학년도"]:
                        grade_idx = idx + 1
                    elif val_norm in ["반", "학급"]:
                        class_idx = idx + 1
                    elif val_norm in ["번호"]:
                        num_idx = idx + 1
                    elif val_norm in ["성명", "이름", "학생명", "학생이름"]:
                        name_idx = idx + 1
                
                # 4개 필수 컬럼이 한 행에 모두 존재하는 경우
                if all(x is not None for x in [grade_idx, class_idx, num_idx, name_idx]):
                    header_row_idx = r
                    col_indices = {"grade": grade_idx, "class": class_idx, "number": num_idx, "name": name_idx}
                    break
            
            if not header_row_idx:
                raise ValueError("나이스 학적현황 양식이 아닙니다. 필수 열(학년, 반, 번호, 성명)을 찾을 수 없습니다.")

            students = []
            
            # 2. 데이터 행 파싱
            for row_cells in ws.iter_rows(min_row=header_row_idx + 1):
                def get_cell_val(col_idx):
                    if col_idx is not None and col_idx <= len(row_cells):
                        return row_cells[col_idx - 1].value
                    return None

                name_val = get_cell_val(col_indices["name"])
                grade_val = get_cell_val(col_indices["grade"])
                class_val = get_cell_val(col_indices["class"])
                num_val = get_cell_val(col_indices["number"])

                if all(x is not None for x in [grade_val, class_val, num_val, name_val]):
                    try:
                        g_num = int(float(str(grade_val).strip()))
                        c_num = int(float(str(class_val).strip()))
                        n_num = int(float(str(num_val).strip()))
                        name_str = str(name_val).strip()
                        
                        if name_str:
                            student_num = f"{g_num}{c_num}{n_num:02d}"
                            students.append(f"{name_str}:학생{student_num}")
                    except (ValueError, TypeError):
                        continue

            return list(dict.fromkeys(students))
        finally:
            wb.close()
