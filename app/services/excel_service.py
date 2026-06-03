import os
import shutil
import uuid
import openpyxl
import pandas as pd
from app.models.detection_model import DetectionItem
from app.utils.logger import logger
from app.utils.path_helper import get_project_root

class ExcelService:
    """
    Excel 파일의 데이터를 읽고 수정하며, 
    저장 시 파일 손상을 막기 위한 Safe Save 메커니즘을 실행하는 비즈니스 서비스 클래스.
    """

    @staticmethod
    def cleanup_temp_folder() -> None:
        """
        앱 시작 시 이전 실행에서 잔류한 temp 파일들을 일괄 정리합니다.

        Safe Save 도중 프로세스가 강제 종료(전원 차단, 크래시 등)되면
        temp/ 폴더에 'temp_' 접두사 파일이 남을 수 있습니다.
        이 메서드는 앱 시작 시 1회 호출하여 해당 파일들을 안전하게 제거합니다.

        - 'temp_' 접두사가 없는 파일은 건드리지 않습니다. (사용자 파일 보호)
        - 삭제 실패는 경고 로그만 남기고 앱 구동을 중단하지 않습니다.
        """
        temp_dir = get_project_root() / "temp"
        if not temp_dir.exists():
            return

        cleaned = 0
        failed = 0
        for entry in temp_dir.iterdir():
            # Safe Save가 생성하는 파일만 선택적으로 삭제 (다른 파일 보호)
            if entry.is_file() and entry.name.startswith("temp_"):
                try:
                    entry.unlink()
                    cleaned += 1
                    logger.debug(f"잔류 temp 파일 삭제: {entry.name}")
                except Exception as e:
                    failed += 1
                    logger.warning(f"잔류 temp 파일 삭제 실패 (무시하고 계속): {entry.name} - {str(e)}")

        if cleaned > 0:
            logger.info(f"앱 시작 시 잔류 temp 파일 {cleaned}개 정리 완료.")
        if failed > 0:
            logger.warning(f"잔류 temp 파일 {failed}개 삭제 실패 (수동 정리 권장).")

    def apply_replacements_safe(self, file_path: str, replacements: list[DetectionItem], output_dir: str) -> str:
        """
        [Safe Save 4단계 적용 구현]
        1. 원본 파일을 읽어서 temp/ 경로에 고유 임시 파일 복사본 생성
        2. openpyxl을 통해 서식을 깨뜨리지 않고 승인된 셀의 데이터를 치환
        3. 임시 경로에 저장 후, 해당 임시 파일이 정상적으로 열리는지 무결성 검증
        4. 무결성 검증 완료 시, 최종 output_dir로 안전하게 이동
        
        :param file_path: 원본 Excel 파일의 경로
        :param replacements: 해당 파일에 적용할 탐지/치환 데이터 목록
        :param output_dir: 최종 저장될 디렉토리 경로
        :return: 최종 저장된 파일의 절대 경로
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"원본 Excel 파일을 찾을 수 없습니다: {file_path}")

        # 1. 임시 파일명 생성 및 복사
        project_root = get_project_root()
        temp_dir = project_root / "temp"
        os.makedirs(temp_dir, exist_ok=True)
        
        file_name = os.path.basename(file_path)
        base_name, ext = os.path.splitext(file_name)
        
        # 고유성 확보를 위해 UUID 기반 임시 파일명 생성
        temp_file_name = f"temp_{uuid.uuid4().hex}{ext}"
        temp_path = str(temp_dir / temp_file_name)
        
        logger.info(f"[Safe Save 1단계] 임시 복사 파일 생성: {temp_path}")
        shutil.copy2(file_path, temp_path)

        wb = None
        try:
            # 2. openpyxl을 통한 서식 보존 셀 수정
            logger.info(f"[Safe Save 2단계] 셀 수정 반영 중: {file_name}")
            wb = openpyxl.load_workbook(temp_path, data_only=False) # 수식 보존을 위해 data_only=False
            
            # 이 파일에 대한 변경 항목 수집
            file_replacements = [r for r in replacements if r.file_path == file_path and r.approved]

            # ⚠️ 연쇄 치환 방지: (시트명, 셀주소) 기준으로 그룹화하여
            # 동일 셀에 대한 모든 치환을 원본 값 기준으로 한 번에 적용.
            # 기존 방식은 항목마다 cell.value를 즉시 덮어써서 다음 항목이
            # 이미 변경된 값을 대상으로 치환하는 연쇄 오염 문제가 있었음.
            cell_groups: dict[tuple[str, str], list] = {}
            for item in file_replacements:
                key = (item.sheet_name, item.cell_address)
                cell_groups.setdefault(key, []).append(item)

            for (sheet_name, cell_address), items in cell_groups.items():
                if sheet_name not in wb.sheetnames:
                    logger.warning(f"존재하지 않는 시트 무시됨: {sheet_name}")
                    continue

                sheet = wb[sheet_name]
                cell = sheet[cell_address]

                # ⚠️ 수식 보호 장치: 수식 셀(첫글자 '=')은 치환 예외
                val_str = str(cell.value or "")
                if val_str.startswith("="):
                    logger.warning(f"수식 포함 셀 치환 방지 적용 (스킵됨): {sheet_name}!{cell_address}")
                    continue

                if cell.value is not None:
                    # 셀 값을 한 번만 읽고, 해당 셀의 모든 치환 항목을 순차 적용 후 한 번만 기록
                    current_val = str(cell.value)
                    new_val = current_val
                    for item in items:
                        new_val = new_val.replace(item.match_value, item.replacement)
                    cell.value = new_val
                    logger.debug(f"수정 완료 - {sheet_name}!{cell_address}: {current_val} -> {new_val}")

            wb.save(temp_path)
            logger.info("[Safe Save 2단계] 임시 수정 저장 성공.")

            # 3. 임시 파일 무결성 검증 (Open Test)
            logger.info("[Safe Save 3단계] 임시 저장 파일 무결성(Open Test) 검증 시작.")
            try:
                test_wb = openpyxl.load_workbook(temp_path, read_only=True)
                test_wb.close()
                logger.info("[Safe Save 3단계] 임시 파일 손상 없음 확인 (검증 완료).")
            except Exception as test_err:
                logger.critical(f"임시 파일 무결성 검증 실패 (파일 손상 가능성): {str(test_err)}")
                raise IOError("수정된 Excel 임시 파일이 정상적으로 로드되지 않습니다. 저장을 중단합니다.") from test_err

            # 4. 최종 목적지로 이동 (원본을 직접 덮어쓰지 않고 output 디렉토리에 새 파일명으로 저장)
            os.makedirs(output_dir, exist_ok=True)
            # 파일명이 중복되지 않도록 접미사 추가 처리 (_anonymized)
            final_file_name = f"{base_name}_anonymized{ext}"
            final_path = os.path.join(output_dir, final_file_name)
            
            # 기존 결과 덮어쓰기 방지: 중복 파일명 존재 시 (1), (2) 등 숫자를 순차적으로 부여
            counter = 1
            while os.path.exists(final_path):
                final_file_name = f"{base_name}_anonymized({counter}){ext}"
                final_path = os.path.join(output_dir, final_file_name)
                counter += 1
                
            logger.info(f"[Safe Save 4단계] 임시 파일을 최종 위치로 이동: {final_path}")
            shutil.move(temp_path, final_path)
            logger.info(f"Safe Save 프로세스 정상 완료: {final_path}")
            return final_path

        except Exception as e:
            logger.error(f"Excel 익명화 중 예외 발생: {str(e)}")
            # 에러 발생 시 임시 파일 클린업 (Recovery 전략)
            if os.path.exists(temp_path):
                try:
                    os.remove(temp_path)
                    logger.info(f"실패 복구 - 임시 파일 삭제 완료: {temp_path}")
                except Exception as clean_err:
                    logger.error(f"실패 복구 중 임시 파일 삭제 실패: {str(clean_err)}")
            raise e
        finally:
            if wb is not None:
                try:
                    wb.close()
                except Exception as close_err:
                    logger.error(f"임시 workbook 자원 해제(close) 실패: {str(close_err)}")

    def save_mapping_file(self, mapping_data: dict[str, str], output_path: str, format_type: str) -> None:
        """
        익명화된 학생 이름/학교명 매핑 관계를 CSV 또는 Excel로 저장합니다.
        예: {'김민수': '학생1', '이서연': '학생2', '서울중학교': '학교A'}
        """
        if not mapping_data:
            logger.info("저장할 매핑 데이터가 비어 있습니다.")
            return

        # Pandas DataFrame으로 정형화
        df = pd.DataFrame(list(mapping_data.items()), columns=["원본 값", "익명화 값"])
        
        try:
            if format_type.upper() == "CSV":
                # CSV로 안전하게 UTF-8-sig 저장 (한글 깨짐 방지)
                df.to_csv(output_path, index=False, encoding="utf-8-sig")
                logger.info(f"매핑 정보가 CSV 파일로 저장되었습니다: {output_path}")
            elif format_type.upper() == "EXCEL":
                df.to_excel(output_path, index=False)
                logger.info(f"매핑 정보가 Excel 파일로 저장되었습니다: {output_path}")
            else:
                raise ValueError(f"지원하지 않는 파일 포맷 형식: {format_type}")
        except Exception as e:
            logger.error(f"매핑 파일 저장 실패: {str(e)}")
            raise e
