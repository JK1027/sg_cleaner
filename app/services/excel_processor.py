import openpyxl
from app.services.base_processor import BaseProcessor, ExtractedTextItem
from app.models.detection_model import DetectionItem
from app.utils.logger import logger

class ExcelProcessor(BaseProcessor):
    """
    Excel (.xlsx) 문서의 데이터 추출 및 안전 치환을 처리하는 구체 프로세서 클래스
    """

    def extract_texts(self, file_path: str) -> list[ExtractedTextItem]:
        """
        Excel 통합 문서에서 텍스트 항목을 추출합니다. (숨김 시트, 병합 셀 및 수식 셀 예외 정책 준수)
        """
        extracted_items = []
        try:
            # 수식 분석이 아닌 최종 표시 텍스트 기준 탐색을 위해 data_only=True로 로드
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=False)
        except Exception as e:
            logger.error(f"Excel 파일 로드 실패 ({file_path}): {str(e)}")
            raise e

        try:
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                
                # ⚠️ 숨김 시트 스킵 처리
                if getattr(sheet, "sheet_state", "visible") != "visible":
                    logger.info(f"숨김 시트 스킵됨: {sheet_name}")
                    continue
                    
                # 행과 열을 순회하며 텍스트 추출
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value is None:
                            continue
                            
                        # ⚠️ 병합 셀 예외 처리: 대표 셀이 아닌 MergedCell은 수정/데이터 조회 불가하므로 스킵
                        if type(cell).__name__ == "MergedCell":
                            continue
                            
                        cell_text = str(cell.value)
                        
                        # ⚠️ 수식은 대상에서 제외
                        if cell_text.startswith("="):
                            continue
                            
                        extracted_items.append(ExtractedTextItem(
                            text=cell_text,
                            location_context=sheet_name,
                            location_detail=cell.coordinate
                        ))
        finally:
            wb.close()

        return extracted_items

    def apply_replacements(self, file_path: str, replacements: list[DetectionItem], temp_path: str) -> None:
        """
        수식을 보존한 상태(data_only=False)로 엑셀 복사본을 수정 후 임시 저장합니다.
        """
        try:
            wb = openpyxl.load_workbook(temp_path, data_only=False)
        except Exception as e:
            logger.error(f"Excel 수정용 임시본 로드 실패 ({temp_path}): {str(e)}")
            raise e

        try:
            # 해당 파일에 대한 승인된 변경 항목 필터링
            file_replacements = [r for r in replacements if r.file_path == file_path and r.approved]

            # ⚠️ 연쇄 치환 방지: (시트명, 셀주소) 기준 그룹화하여 원본값 기준으로 한번에 적용
            cell_groups: dict[tuple[str, str], list[DetectionItem]] = {}
            for item in file_replacements:
                key = (item.location_context, item.location_detail)
                cell_groups.setdefault(key, []).append(item)

            for (sheet_name, cell_address), items in cell_groups.items():
                if sheet_name not in wb.sheetnames:
                    logger.warning(f"존재하지 않는 시트 무시됨: {sheet_name}")
                    continue

                sheet = wb[sheet_name]
                cell = sheet[cell_address]

                # ⚠️ 수식 보호 장치
                val_str = str(cell.value or "")
                if val_str.startswith("="):
                    logger.warning(f"수식 포함 셀 치환 방지 적용 (스킵됨): {sheet_name}!{cell_address}")
                    continue

                if cell.value is not None:
                    current_val = str(cell.value)
                    new_val = current_val
                    for item in items:
                        new_val = new_val.replace(item.match_value, item.replacement)
                    cell.value = new_val
                    logger.debug(f"수정 완료 - {sheet_name}!{cell_address}: {current_val} -> {new_val}")

            wb.save(temp_path)
            logger.info(f"ExcelProcessor: 임시 수정 반영 완료. ({temp_path})")
        finally:
            wb.close()
