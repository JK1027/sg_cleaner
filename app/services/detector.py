import openpyxl
from app.models.detection_model import DetectionItem
from app.utils.logger import logger


def _make_school_label(n: int) -> str:
    """
    1 → 'A', 2 → 'B', ..., 26 → 'Z', 27 → 'AA', 28 → 'AB', ...
    Excel 컬럼 명명 방식과 동일하게 무제한 확장 가능한 알파벳 레이블 생성.
    :param n: 1-based 인덱스
    :return: 알파벳 레이블 문자열
    """
    result = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        result = chr(65 + remainder) + result
    return result

class AnonymizeDetector:
    """
    사용자가 입력한 학생 이름/학교명 패턴 목록에 근거하여 Excel 셀의 텍스트 데이터를 검색 및 탐지하는 엔진.
    """
    def __init__(self, student_names: list[str], school_names: list[str],
                 delete_keywords: list[str] = None, delete_replacement: str = ""):
        # 공백 제거 및 중복 제거 처리 (순서 보존)
        self.student_names = list(dict.fromkeys([name.strip() for name in student_names if name.strip()]))
        self.school_names = list(dict.fromkeys([school.strip() for school in school_names if school.strip()]))
        self.delete_keywords = list(dict.fromkeys([word.strip() for word in (delete_keywords or []) if word.strip()]))
        self.delete_replacement = delete_replacement
        
        # 탐색 과정 중 고유한 익명화 대체 이름(예: 학생1, 학교A)을 관리하기 위한 매핑 저장소
        self.student_mapping = {} # {'김민수': '학생1'}
        self.school_mapping = {}   # {'서울중학교': '학교A'}
        self.delete_mapping = {}   # {'지울단어': ''}

    def scan_workbook(self, file_path: str) -> list[DetectionItem]:
        """
        주어진 Excel 통합 문서를 읽고, 모든 활성 시트의 셀을 탐색하여 매칭 결과를 수집합니다.
        
        :param file_path: 탐색 대상 Excel 파일 경로
        :return: DetectionItem 객체 목록
        """
        results = []
        try:
            # 스캔 시에는 수식 분석이 아닌 최종 표시 텍스트를 기준으로 탐색해야 하므로
            # data_only=True로 로드합니다. (수정 시에는 data_only=False 사용)
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=False)
        except Exception as e:
            logger.error(f"Excel 파일 로드 실패 ({file_path}): {str(e)}")
            raise e

        # 매핑 카운터
        student_count = len(self.student_mapping) + 1
        school_count = len(self.school_mapping) + 1

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            # ⚠️ 숨김 시트 스킵 처리
            if getattr(sheet, "sheet_state", "visible") != "visible":
                logger.info(f"숨김 시트 스킵됨: {sheet_name}")
                continue
                
            logger.info(f"시트 스캔 중: {sheet_name} ({file_path})")
            
            # 행과 열을 순회하며 데이터 탐색
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                        
                    # ⚠️ 병합 셀 예외 처리:
                    # openpyxl에서 대표 셀이 아닌 하위 병합 셀들은 MergedCell 타입으로 분류됨.
                    # MergedCell은 수정 및 데이터 조회가 불가능하므로 스킵 처리.
                    if type(cell).__name__ == "MergedCell":
                        continue
                        
                    cell_text = str(cell.value)
                    
                    # ⚠️ 수식은 탐지 및 수정 대상에서 제외
                    if cell_text.startswith("="):
                        continue
                        
                    # 모든 검색 패턴 통합 및 길이 역순 정렬
                    patterns = []
                    for name in self.student_names:
                        patterns.append((name, "student"))
                    for school in self.school_names:
                        patterns.append((school, "school"))
                    for word in self.delete_keywords:
                        patterns.append((word, "delete"))
                        
                    patterns.sort(key=lambda x: len(x[0]), reverse=True)
                    
                    matched_intervals = [] # list of (start_idx, end_idx)
                    
                    for pattern, pat_type in patterns:
                        start_find = 0
                        while True:
                            idx = cell_text.find(pattern, start_find)
                            if idx == -1:
                                break
                            
                            pat_len = len(pattern)
                            end_idx = idx + pat_len
                            
                            # 기존에 매칭된 구간과 겹치는지 체크
                            overlap = False
                            for m_start, m_end in matched_intervals:
                                if not (end_idx <= m_start or idx >= m_end):
                                    overlap = True
                                    break
                                    
                            if not overlap:
                                matched_intervals.append((idx, end_idx))
                                
                                # 매칭 데이터 생성
                                if pat_type == "student":
                                    if pattern not in self.student_mapping:
                                        self.student_mapping[pattern] = f"학생{student_count}"
                                        student_count += 1
                                    item = DetectionItem(
                                        file_path=file_path,
                                        sheet_name=sheet_name,
                                        cell_address=cell.coordinate,
                                        original_value=cell_text,
                                        match_value=pattern,
                                        replacement=self.student_mapping[pattern],
                                        approved=True
                                    )
                                    results.append(item)
                                    logger.debug(f"이름 탐지: {cell.coordinate} -> {pattern}")
                                    
                                elif pat_type == "school":
                                    if pattern not in self.school_mapping:
                                        self.school_mapping[pattern] = f"학교{_make_school_label(school_count)}"
                                        school_count += 1
                                    item = DetectionItem(
                                        file_path=file_path,
                                        sheet_name=sheet_name,
                                        cell_address=cell.coordinate,
                                        original_value=cell_text,
                                        match_value=pattern,
                                        replacement=self.school_mapping[pattern],
                                        approved=True
                                    )
                                    results.append(item)
                                    logger.debug(f"학교명 탐지: {cell.coordinate} -> {pattern}")
                                    
                                elif pat_type == "delete":
                                    if pattern not in self.delete_mapping:
                                        self.delete_mapping[pattern] = self.delete_replacement
                                    item = DetectionItem(
                                        file_path=file_path,
                                        sheet_name=sheet_name,
                                        cell_address=cell.coordinate,
                                        original_value=cell_text,
                                        match_value=pattern,
                                        replacement=self.delete_mapping[pattern],
                                        approved=True
                                    )
                                    results.append(item)
                                    logger.debug(f"삭제 단어 탐지: {cell.coordinate} -> {pattern}")
                                    
                            # 해당 위치 다음부터 검색 지속
                            start_find = idx + 1

        wb.close()
        logger.info(f"스캔 완료: {file_path} (탐색 개수: {len(results)}개)")
        return results

    def get_full_mapping(self) -> dict[str, str]:
        """학생, 학교 및 삭제 키워드의 전체 누적 치환 매핑 대장을 병합하여 반환합니다."""
        full_map = {}
        full_map.update(self.student_mapping)
        full_map.update(self.school_mapping)
        full_map.update(self.delete_mapping)
        return full_map
